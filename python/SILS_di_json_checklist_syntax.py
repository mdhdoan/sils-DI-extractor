# import libraries 
import ast
from datetime import datetime
import json
import os 
import random
import shutil
import sys

from langchain.output_parsers.list import NumberedListOutputParser

from openpyxl import load_workbook, Workbook

# from langchain.chains.llm import LLMChain
from langchain_ollama import OllamaLLM
from langchain_core.prompts import PromptTemplate
# from langchain.prompts.prompt import PromptTemplate

from azure.storage.blob import BlobServiceClient

# LLM model
model_select = 'llama3.1'

random.seed(2024)
file_list_in_directory = os.listdir(sys.argv[1])
format = sys.argv[3]
worksheet = load_workbook(sys.argv[2])[format]

# current_directory = os.getcwd()
# connection_file = os.path.join(current_directory, "connections.json")
# connection_details = {}
# with open(connection_file, "r+") as connection_data:
#     connection_details = json.load(connection_data)

output_azure = False
output_connection_string = ""
output_success_container_name = "quesnel-ground/success_jsons" 
output_update_container_name = "quesnel-ground/update_jsons" 
output_flagged_container_name = "quesnel-ground/flagged_jsons" 
output_report_container_name =  "quesnel-ground"

error_list = ['Ø', 'ø']
survey_list = ['Survey Type: Ground','Survey Type: Aerial']
water_level_list = ['Water Level: % Bankfill: <25%',
                    'Water Level: % Bankfill: 25-50%','Water Level: % Bankfill: 50-75%',
                    'Water Level: % Bankfill: 75-100%','Water Level: % Bankfill: +100%']
weather_brightness_list = ['Weather: Brightness: Full','Weather: Brightness: Bright','Weather: Brightness: Medium',
                    'Weather: Brightness: Dark']
weather_cloudy_list = ['Weather: %Cloudy: 0%','Weather: %Cloudy: 25%',
                    'Weather: %Cloudy: 50%','Weather: %Cloudy: 75%','Weather: %Cloudy: 100%']
precipitation_type_list = ['Precipitation: Type: Rain','Precipitation: Type: Snow','Precipitation: Type: None']
precipitation_intensity_list = [
                    'Precipitation: Intensity: Light','Precipitation: Intensity: Medium',
                    'Precipitation: Intensity: Heavy']
fish_visibility_list = ['Water Conditions: Fish Visibility: Low',
                    'Water Conditions: Fish Visibility: Medium','Water Conditions: Fish Visibility: High']
water_clarity_list = ['Water Conditions: Water Clarity: 0-0.25m','Water Conditions: Water Clarity: 0.25-0.5m',
                    'Water Conditions: Water Clarity: 0.5-1.0m','Water Conditions: Water Clarity: 1-3m',
                    'Water Conditions: Water Clarity: 3m to bottom']


output_parser = NumberedListOutputParser()  
format_instructions = output_parser.get_format_instructions()

llm = OllamaLLM(model = model_select, temperature = 0.0)

def create_prompt(format_instructions):
    QA_TEMPLATE = """
        Format as ['True', '<reason>'] if true, ['False', '<reason>'] if false
        Only return one list result please. Check if your answer returns a list formatted like above before returning. Answer: {format_instructions}
        Is {value} best described by given {description} AND follows syntaxtically of {example}?
    """
    #Verify your answer, and if the result list has more than 2 items, then Value has multiple parts. Treat them all as one value only, and ignore the number in brackets in them. Retry to shorten it to format above.
    return PromptTemplate(
        input_variables=["value", "description", "example"], 
        partial_variables={"format_instructions": format_instructions},
        template=QA_TEMPLATE)

prompt = create_prompt(format_instructions)

llm_chain = prompt | llm | output_parser

def llm_check_value(text_content, definition, example):
    print('VALUE:', text_content, flush = True, end = '\t')
    # LLM_start_time = datetime.now()
    # print(text_content, definition, example, flush = True)
    result = llm_chain.invoke({'value': text_content, 'description': definition, 'example': example})
    print(result, '--- becomes --->', flush = True, end = ' ')
    result_sample = ast.literal_eval(result[0])
    # LLM_end_time = datetime.now()
    ### Test if result is a valid list to be considered
    if type(result) is not list or result == []:
        result = ['False', 'QA gave bad result']
        print(result, flush = True)
        return result
    if type(result[0]) is not str:
        print(result[0], 'is not a string, sus on file', flush = True)
        result = ['False', 'QA gave bad result']
        print(result, flush = True)
        return result
    ### Result here is now understand as a list of strings
    if len(result) > 2:
        result = result[0:2] 
        print('RESULT LONG LIST', result, flush = True)

    ### Result here is now considered as a list of two elements or less
    elif len(result) < 2:
        if type(result_sample) is not list:
            result = ['False', result[0]]
            print('RESULT NOT LIST', result, flush = True)
            return result
        if result[0] not in ['False', 'True']: ### This is for list of no boolean value
            result = ['False', result[0]]
            print('SHORT RESULT NOT BOOL', result, flush = True)
            if type(eval(result[1])) is list:
                result = ast.literal_eval(result[1])
                print('RESULT LIST IN LIST', type(result), '-', result, flush = True)
                return result
            return result
        if eval(result[0]): ### this is for list of one element ['True'] or ['False']
            result = [result[0], 'No explanation']
            print(result, flush = True)
        else:            
            result = [result[0], 'Value does not match provided standards']
            print(result, flush = True)
   
    ### Result here is now considered as a list of two elements precisely
    if result[0] not in ['False', 'True']:
        result = ['False', result[0]]
        print('RESULT NOT BOOL', result, flush = True)
        return result
    if type(eval(result[0])) is list:
        result = ast.literal_eval(result[0])
        print('RESULT LIST IN LIST', type(result), '-', result, flush = True)
        return result

    print('RESULT ALL GOOD', flush = True)
    return result

def load_worksheet_as_dict():
    result_dict = {}
    for row_i, row in enumerate(worksheet['D'], 1):
        # print(row)
        # print(worksheet.cell(row_i, column = 4).value)
        if row.value is not None:
            result_dict[row.value.lower().replace(':', '')] = row_i
    return result_dict

worksheet_dict = load_worksheet_as_dict()
# print(worksheet_dict)

def check_field(field_name):
    # print(field_name.replace(':', ''), end = '\t')
    row_i = worksheet_dict.get(field_name.lower().replace(':', '').replace('Visability', 'Visibility'))
    # print(row_i)
    if row_i is None:
        return ['No definition available', 'NA']
    return [
        worksheet.cell(row=row_i, column=6).value, 
        worksheet.cell(row=row_i, column=7).value
    ]

def clean_value(value):
    replace_with_space = ['\n', '*']
    replace_with_empty = [':', '\u00b0', '\u2103', '\u00d3', '\"', '\u00b7']
    for char in replace_with_space:
        value = value.replace(char, ' ')
    for char in replace_with_empty:
        value = value.replace(char, '')
    # print('fixed value:', value, flush = True)
    return value

def match_value_to_llm_qa(value, definition, example):
    check, standard = llm_check_value(value, definition, example)
    # print('\tcheck:', check, '\t|\tstandard:', standard)
    # print(".", end = '', flush = True)
    if eval(check):
        # print(f"{content} PASSED\n", flush = True) #:\t{content}
        return 'matched'
    else:
        return standard
        # print(f"<{content}> FAILED\n\t{standard}\n", flush = True)

def check_unique(json_dict, field_list):
    value_list = []
    for field in field_list:
        value_list += json_dict[field][0]
    if value_list.count('selected') > 1:
        return 'Multi-checkmarks detected'
    else:
        return 'All unique'

def process_json(json_file, json_uuid):
    json_error_dict = {}
    total_field_list = [*survey_list, *water_level_list, *weather_brightness_list, *weather_cloudy_list, *precipitation_type_list, 
                        *precipitation_intensity_list, *fish_visibility_list, *water_clarity_list]
    list_of_total_field_list = [survey_list, water_level_list, weather_brightness_list, weather_cloudy_list, precipitation_type_list, 
                        precipitation_intensity_list, fish_visibility_list, water_clarity_list]
    for field, value in json_file.items():
        # print(f"{field}\t:\t{value}", flush = True)
        if field in total_field_list:
            value[0] = clean_value(value[0])
            if value[0] not in['selected', 'unselected']:
                json_error_dict[field] = 'Data is in wrong format'
            else:
                for field_list in list_of_total_field_list:
                    unique_valid = check_unique(json_file, field_list)
                    if unique_valid != 'All unique':
                        json_error_dict[field] = unique_valid
        elif field in ['SK_Count_data', 'SK_Count_data_2', 'SK_Count_data_3']:
            for row in value:
                for column, cell in row.items():
                    # print(column, '-', cell)
                    if cell[0] in error_list:
                        cell[0] = '0'
                        # print('\t', column, '-', cell)
                        definition, example = check_field(column)
                        if definition == 'No definition available':
                            continue
                        else:
                            cell[0] = clean_value(cell[0])
                            content = cell[0]
                            check_value = match_value_to_llm_qa(cell[0], definition, example)
                            if check_value == 'matched':
                                continue
                            else: 
                                json_error_dict[field][column] = check_value
                # if json_file['SK_Count_data_2']['Female'] != 0 and 
        else:
            # print(field)
            definition, example = check_field(field)
            value[0] = clean_value(value[0])
            content = value[0]
            if definition == 'NA' or content is None:
                print('content', content)
                continue
            else:
                check_value = match_value_to_llm_qa(content, definition, example)
                if check_value == 'matched':
                    continue
                else: 
                    json_error_dict[field] = check_value
    json_new_data = json.dumps(json_file, indent = 4)
    # print(json_new_data)
    file_address = os.path.join(sys.argv[1], json_uuid + '.json')
    # print(file_address)
    with open(file_address, 'w+') as update_json_file:
        update_json_file.write(json_new_data)
    print(f"\nQA Completed", flush = True)
    return json_error_dict

# function writes to Azure, push data
def write_to_azure(file_name, output_file, container_name):
    blob_service_client = BlobServiceClient.from_connection_string(output_connection_string)
    container_client = blob_service_client.get_container_client(container_name)
    with open(output_file, "rb") as data:
        container_client.upload_blob(name=file_name, data=data, overwrite=True)

def write_to_file(file_list_in_directory, error_dict):
    for json_file in file_list_in_directory:
        if json_file.endswith('.json'):
            json_uuid = json_file[:-5]
            file_address = os.path.join(sys.argv[1], json_file)
            if json_uuid not in error_dict:
                print(f"{json_uuid} not found in error_dict", flush = True)
                continue
            destination_file = './'+ output_report_container_name + '_checked_json' + '/' + json_file
            if not os.path.isdir('./'+ output_report_container_name + '_checked_json' + '/'):
                os.mkdir('./'+ output_report_container_name + '_checked_json' )
            # print(f"{json_uuid} passed", flush = True, end = '')
            shutil.copy(file_address, destination_file)
            if output_azure:
                print('\t|\tUploading to storage account...', flush = True, end = '')
                write_to_azure(json_file, file_address, output_success_container_name)
                write_to_azure(json_file, file_address, output_update_container_name)
                print('done')
    report_name = output_report_container_name + '_report.json'
    # report_name = model_select + '_report.json'
    json_error_data = json.dumps(error_dict, indent=4)
    with open(report_name, 'w') as report_file:
        report_file.write(json_error_data)
    print('done')

def check_json_syntax_local():
    start_time = datetime.now()
    loaded_list_of_json_files = {}
    for json_file in file_list_in_directory:
        if json_file.endswith('.json'):
            file_address = os.path.join(sys.argv[1], json_file)
            uuid = json_file[:-5]
            # print(uuid)
            with open(file_address, 'rt', encoding='utf-8') as file:
                doc = json.load(file)
                loaded_list_of_json_files[uuid] = doc
        else:
            continue
    error_dict = {}
    for uuid, json_file in loaded_list_of_json_files.items():
        print(uuid)
        error_dict[uuid] = {'error_score': 1, 'problem': 'no QA available'}
        try:
            error_dict[uuid] = process_json(json_file, uuid)
            error_dict[uuid]['error_score'] = len(error_dict[uuid])
            print('error_score:', error_dict[uuid]['error_score'], flush = True)
            ### error_dict[2019.json]: {'Date': [], 'error_score' = 0}###
            # print(error_dict)
        except:
            print('error hit for', uuid)
            continue
    # print(error_dict)
    write_to_file(file_list_in_directory, error_dict)

    end_time = datetime.now()
    seconds = (end_time - start_time).total_seconds()
    print(f"Total Execution time: {seconds} secs for {len(file_list_in_directory)} files at {end_time}", flush=True)
        
###---------------------------------------------------------------###
if __name__ == "__main__": 
    print("Running local at", datetime.now()) 
    check_json_syntax_local()

## Example: python .\python\di_json_checklist_syntax.py .\json\sils-ground\raw_jsons\ .\Data_Standards\Data_Standard_Quesnel_Roving.xlsx 'Data_Standard_Quesnel_Roving'
# python3 .\python\SILS_di_json_checklist_syntax.py .\test-data\ .\standards\Data_Standard_Quesnel_Roving.xlsx 'Data_Standard_Quesnel_Roving'